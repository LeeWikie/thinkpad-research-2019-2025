#!/usr/bin/env python3
"""ThinkPad调研项目 - 最终Excel生成脚本 (Task 21)

基于合并数据生成包含条件格式、数据说明Sheet的最终Excel文件。
"""

import json
import re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, NamedStyle, numbers
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── 路径配置 ──────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUT_PATH = PROJECT_ROOT / "data" / "processed" / "merged_data.json"
OUTPUT_PATH = PROJECT_ROOT / "thinkpad_complete_research.xlsx"


# ── 样式常量 ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Microsoft YaHei", size=10)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=False)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Linux兼容性评分 条件格式颜色（星级数→填充色）
RATING_COLORS = {
    "★★★★★": "00B050",  # 5星=绿色
    "★★★★☆": "92D050",  # 4星=浅绿
    "★★★☆☆": "FFFF00",  # 3星=黄色
    "★★☆☆☆": "FFC000",  # 2星=橙色
    "★☆☆☆☆": "FF0000",  # 1星=红色（字体白色）
    "N/A":     "C0C0C0",  # 灰色
}

RATING_FONT_COLORS = {
    "★★★★★": "FFFFFF",
    "★★★★☆": "000000",
    "★★★☆☆": "000000",
    "★★☆☆☆": "000000",
    "★☆☆☆☆": "FFFFFF",
    "N/A":     "000000",
}

# Ubuntu认证颜色
UBUNTU_COLORS = {
    "是":     "00B050",
    "否":     "FF0000",
    "未确认": "FFFF00",
}
UBUNTU_FONT_COLORS = {
    "是": "FFFFFF",
    "否": "FFFFFF",
    "未确认": "000000",
}


# ── 数据扁平化 ────────────────────────────────────────────

def flatten_list(value, sep=", "):
    """将列表转为字符串。"""
    if isinstance(value, list):
        return sep.join(str(v) for v in value if v)
    return str(value) if value is not None else ""


def flatten_feedback(feedback_list):
    """将社区反馈列表转为可读字符串。"""
    if not feedback_list or not isinstance(feedback_list, list):
        return ""
    parts = []
    for item in feedback_list:
        if isinstance(item, dict):
            src = item.get("source", "未知")
            content = item.get("content", "")
            parts.append(f"[{src}] {content}")
        else:
            parts.append(str(item))
    return " | ".join(parts)


def flatten_breakdown(breakdown_dict):
    """将评分细分字典转为字符串。"""
    if not breakdown_dict or not isinstance(breakdown_dict, dict):
        return ""
    return "; ".join(f"{k}: {v}" for k, v in breakdown_dict.items())


def prepare_dataframe(models):
    """将原始模型数据展平为DataFrame。"""
    records = []
    for m in models:
        record = {
            "型号": m.get("model", ""),
            "系列": m.get("series", ""),
            "发布年份": m.get("year", ""),
            "代际": m.get("generation", ""),
            "尺寸(英寸)": m.get("screen_size", ""),
            "分辨率": m.get("resolution", ""),
            "刷新率(Hz)": m.get("refresh_rate", ""),
            "触摸屏": m.get("touch_screen", ""),
            "CPU型号": m.get("cpu", ""),
            "CPU架构": m.get("cpu_architecture", ""),
            "集成显卡": m.get("igpu", ""),
            "独立显卡": m.get("dgpu", ""),
            "内存类型": m.get("ram_type", ""),
            "最大内存": m.get("ram_max", ""),
            "存储接口": m.get("storage", ""),
            "接口": m.get("ports", ""),
            "无线": m.get("wireless", ""),
            "指纹识别": m.get("fingerprint", ""),
            "摄像头": m.get("camera", ""),
            "官方标称续航(h)": m.get("battery_life_official", ""),
            "重量": m.get("weight", ""),
            "Ubuntu认证": "是" if m.get("ubuntu_certified") is True else "否",
            "Ubuntu认证版本": flatten_list(m.get("ubuntu_versions")),
            "Ubuntu数据来源": flatten_list(m.get("ubuntu_source")),
            "Fedora兼容性": m.get("fedora_compatibility", ""),
            "Fedora认证版本": flatten_list(m.get("fedora_versions")),
            "Fedora说明": m.get("fedora_notes", ""),
            "Fedora已知问题": flatten_list(m.get("fedora_known_issues")),
            "Arch兼容性": m.get("arch_compatibility", ""),
            "Arch已知问题": flatten_list(m.get("arch_known_issues")),
            "Arch Wiki页面": m.get("arch_wiki_page", ""),
            "社区评分": m.get("community_rating", ""),
            "社区正面反馈": flatten_feedback(m.get("community_positive_feedback")),
            "社区负面反馈": flatten_feedback(m.get("community_negative_feedback")),
            "二手价格区间(CNY)": (
                f"{m['price_range_cny'][0]}-{m['price_range_cny'][1]}"
                if isinstance(m.get("price_range_cny"), list) and len(m["price_range_cny"]) == 2
                else ""
            ),
            "典型配置": m.get("price_typical_config", ""),
            "成色": m.get("price_condition", ""),
            "价格数据质量": m.get("price_data_quality", ""),
            "Linux兼容性原始分数": m.get("linux_compat_score_raw", ""),
            "Linux兼容性评分": m.get("linux_compat_rating", ""),
            "Linux兼容性细分": flatten_breakdown(m.get("linux_compat_breakdown")),
            "数据可信度": m.get("data_credibility", ""),
            "数据来源URL": m.get("data_source_urls", ""),
            "数据更新日期": m.get("data_update_date", ""),
        }
        records.append(record)
    return pd.DataFrame(records)


# ── 条件格式与样式 ─────────────────────────────────────────

def apply_header_style(ws, num_cols):
    """设置标题行样式。"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_data_style(ws, num_rows, num_cols):
    """设置数据区域样式。"""
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = border
            cell.alignment = DATA_ALIGNMENT


def apply_conditional_formatting(ws, df):
    """应用条件格式：Linux兼容性评分 + Ubuntu认证。"""
    num_rows = len(df) + 1  # +1 for header

    # 定位关键列（通过标题行查找）
    rating_col = None
    ubuntu_col = None
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col).value
        if header_val == "Linux兼容性评分":
            rating_col = col
        elif header_val == "Ubuntu认证":
            ubuntu_col = col

    # ── Linux兼容性评分条件格式 ──
    if rating_col:
        col_letter = get_column_letter(rating_col)
        for rating, color in RATING_COLORS.items():
            font_color = RATING_FONT_COLORS[rating]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            font = Font(name="Microsoft YaHei", size=10, bold=True,
                       color=font_color)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{num_rows}",
                CellIsRule(
                    operator="equal",
                    formula=[f'"{rating}"'],
                    fill=fill,
                    font=font,
                )
            )

    # ── Ubuntu认证条件格式 ──
    if ubuntu_col:
        col_letter = get_column_letter(ubuntu_col)
        for cert_value, color in UBUNTU_COLORS.items():
            font_color = UBUNTU_FONT_COLORS[cert_value]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            font = Font(name="Microsoft YaHei", size=10, bold=True,
                       color=font_color)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{num_rows}",
                CellIsRule(
                    operator="equal",
                    formula=[f'"{cert_value}"'],
                    fill=fill,
                    font=font,
                )
            )

    # ── Fedora兼容性 条件格式 ──
    fedora_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Fedora兼容性":
            fedora_col = col
            break
    if fedora_col:
        col_letter = get_column_letter(fedora_col)
        fedora_colors = {
            "certified":     ("00B050", "FFFFFF"),
            "ships_fedora":  ("00B050", "FFFFFF"),
            "friendly":      ("92D050", "000000"),
            "community_good":  ("C6EFCE", "000000"),
            "community_warning": ("FFC000", "000000"),
            "unknown":       ("C0C0C0", "000000"),
            "N/A":           ("C0C0C0", "000000"),
        }
        for fval, (bg, fg) in fedora_colors.items():
            fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            font = Font(name="Microsoft YaHei", size=10, bold=True, color=fg)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{num_rows}",
                CellIsRule(operator="equal", formula=[f'"{fval}"'],
                          fill=fill, font=font)
            )

    # ── 数据可信度 条件格式 ──
    cred_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "数据可信度":
            cred_col = col
            break
    if cred_col:
        col_letter = get_column_letter(cred_col)
        cred_colors = {
            "中高 (2个官方/权威来源)":  ("00B050", "FFFFFF"),
            "中 (1个官方来源)":        ("92D050", "000000"),
            "低 (无官方来源)":         ("FFC000", "000000"),
        }
        for cval, (bg, fg) in cred_colors.items():
            fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            font = Font(name="Microsoft YaHei", size=10, bold=True, color=fg)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{num_rows}",
                CellIsRule(operator="equal", formula=[f'"{cval}"'],
                          fill=fill, font=font)
            )


def auto_width(ws, max_width=40, min_width=8):
    """自适应列宽。"""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                # 中文字符算2个宽度
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                if length > max_len:
                    max_len = length
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted

    # 文本较长列额外加宽
    wide_cols = {
        "社区正面反馈": 50,
        "社区负面反馈": 50,
        "Fedora说明": 40,
        "Fedora已知问题": 35,
        "接口": 30,
        "典型配置": 30,
        "Linux兼容性细分": 30,
        "Ubuntu数据来源": 50,
        "Arch Wiki页面": 50,
    }
    for col_cells in ws.columns:
        header = col_cells[0].value
        if header in wide_cols:
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = wide_cols[header]


# ── 数据说明Sheet ──────────────────────────────────────────

def create_data_description_sheet(wb):
    """创建数据说明Sheet。"""
    ws = wb.create_sheet("数据说明")

    # 标题
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "ThinkPad Linux兼容性调研 - 数据说明"
    title_cell.font = Font(name="Microsoft YaHei", bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 基本信息
    info = [
        ("数据生成日期", "2026-06-08"),
        ("总型号数", "162"),
        ("覆盖年份", "2019-2025"),
        ("涉及系列", "T, X1 Carbon/Nano/Yoga/Titanium/Fold, X, P, L, E, Z"),
    ]
    row = 3
    ws.cell(row=row, column=1, value="── 基本信息 ──").font = Font(
        name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
    row += 1
    for label, value in info:
        ws.cell(row=row, column=1, value=label).font = Font(
            name="Microsoft YaHei", bold=True, size=10)
        ws.cell(row=row, column=2, value=value).font = Font(
            name="Microsoft YaHei", size=10)
        row += 1

    # 字段说明
    row += 1
    ws.cell(row=row, column=1, value="── 字段说明 ──").font = Font(
        name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
    row += 1

    field_descriptions = [
        ("型号", "ThinkPad完整型号名称"),
        ("系列", "所属产品系列 (T Series/X1 Carbon/E Series等)"),
        ("发布年份", "产品发布年份"),
        ("代际", "产品代际标识"),
        ("尺寸(英寸)", "屏幕尺寸"),
        ("分辨率", "屏幕分辨率 (例: 1920x1200)"),
        ("刷新率(Hz)", "屏幕刷新率"),
        ("触摸屏", "是否支持触摸屏"),
        ("CPU型号", "处理器型号"),
        ("CPU架构", "CPU架构 (Intel/AMD)"),
        ("集成显卡", "集成GPU型号"),
        ("独立显卡", "独立GPU型号 (如无则为'无')"),
        ("内存类型", "内存规格 (DDR4/DDR5/LPDDR4x/LPDDR5x)"),
        ("最大内存", "最大支持内存容量"),
        ("存储接口", "存储接口类型 (NVMe/SATA)"),
        ("接口", "物理接口列表"),
        ("无线", "无线网络规格 (WiFi 6E/WiFi 7等)"),
        ("指纹识别", "是否配备指纹识别模块"),
        ("摄像头", "摄像头规格"),
        ("官方标称续航(h)", "官方标称电池续航时间(小时)"),
        ("重量", "机身重量"),
        ("Ubuntu认证", "是否获得Ubuntu官方认证 (是/否)"),
        ("Ubuntu认证版本", "认证通过的Ubuntu版本列表"),
        ("Ubuntu数据来源", "Ubuntu认证数据来源URL"),
        ("Fedora兼容性", "Fedora兼容性级别 (certified/ships_fedora/friendly/community_good/community_warning/unknown)"),
        ("Fedora认证版本", "兼容的Fedora版本列表"),
        ("Fedora说明", "Fedora兼容性详细说明"),
        ("Fedora已知问题", "Fedora下已知的兼容性问题"),
        ("Arch兼容性", "Arch Linux兼容性状态"),
        ("Arch已知问题", "Arch Linux下已知问题"),
        ("Arch Wiki页面", "Arch Wiki相关页面URL"),
        ("社区评分", "社区综合评分及评级"),
        ("社区正面反馈", "来自社区论坛/博客的正面反馈"),
        ("社区负面反馈", "来自社区论坛/博客的负面反馈"),
        ("二手价格区间(CNY)", "中国二手市场价格范围(人民币)"),
        ("典型配置", "该价格对应的典型配置"),
        ("成色", "二手设备成色描述"),
        ("价格数据质量", "价格数据的可靠性评级"),
        ("Linux兼容性原始分数", "Linux兼容性评分原始数值(0-10)"),
        ("Linux兼容性评分", "Linux兼容性星级评分(★☆☆☆☆ 至 ★★★★★)"),
        ("Linux兼容性细分", "Linux兼容性评分构成细分(各维度得分)"),
        ("数据可信度", "整体数据可信度评级"),
        ("数据来源URL", "原始数据来源URL汇总"),
        ("数据更新日期", "数据最后更新日期"),
    ]

    # 写字段说明表头
    headers = ["字段名", "说明"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.border = THIN_BORDER
    row += 1

    for fname, fdesc in field_descriptions:
        ws.cell(row=row, column=1, value=fname).font = Font(
            name="Microsoft YaHei", bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=fdesc).font = Font(
            name="Microsoft YaHei", size=10)
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # 数据来源汇总
    row += 2
    ws.cell(row=row, column=1, value="── 数据来源汇总 ──").font = Font(
        name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
    row += 1

    sources = [
        ("Lenovo PSREF", "Lenovo官方产品规格数据库", "https://psref.lenovo.com/"),
        ("Ubuntu Certification", "Ubuntu官方硬件认证数据库", "https://ubuntu.com/certified"),
        ("Lenovo Linux Support", "Lenovo官方Linux兼容性矩阵", "https://support.lenovo.com/solutions/pd031426"),
        ("Fedora Discussion", "Fedora社区ThinkPad兼容性列表", "https://discussion.fedoraproject.org/"),
        ("Arch Wiki", "Arch Linux官方Wiki - Laptop/Lenovo页面", "https://wiki.archlinux.org/title/Laptop/Lenovo"),
        ("专门网 (ibmnb.com)", "中国ThinkPad专业论坛二手交易区", "https://www.ibmnb.com/"),
        ("ibmbjb", "ThinkPad二手交易信息平台", "https://www.ibmbjb.com/"),
        ("Reddit r/thinkpad", "Reddit ThinkPad社区讨论", "https://www.reddit.com/r/thinkpad/"),
        ("Reddit r/linuxonthinkpad", "Reddit ThinkPad Linux专题讨论", "https://www.reddit.com/r/linuxonthinkpad/"),
        ("V2EX", "中国技术社区ThinkPad相关讨论", "https://www.v2ex.com/"),
        ("知乎", "中国问答平台ThinkPad相关讨论", "https://www.zhihu.com/"),
        ("Linux-Hardware.org", "开源硬件探测数据库", "https://linux-hardware.org/"),
    ]

    src_headers = ["数据来源", "说明", "URL"]
    for i, h in enumerate(src_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.border = THIN_BORDER
    row += 1

    for name, desc, url in sources:
        for i, val in enumerate([name, desc, url], 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.border = THIN_BORDER
            if i == 1:
                cell.font = Font(name="Microsoft YaHei", bold=True, size=10)
        row += 1

    # 评分标准说明
    row += 2
    ws.cell(row=row, column=1, value="── Linux兼容性评分标准 ──").font = Font(
        name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
    row += 1

    rating_criteria = [
        ("★★★★★ (5星)", "完美兼容", "Ubuntu/Fedora/Arch三大发行版均有官方认证或社区确认全功能正常，所有硬件驱动完整，无已知严重问题"),
        ("★★★★☆ (4星)", "优秀兼容", "2个以上发行版获认证或社区确认兼容，绝大部分硬件正常工作，仅少数非关键设备需要手动配置"),
        ("★★★☆☆ (3星)", "良好兼容", "至少1个发行版获认证或确认兼容，核心功能(显示/网络/音频)正常，部分设备(指纹/NVIDIA)需额外驱动"),
        ("★★☆☆☆ (2星)", "部分兼容", "仅社区确认基本兼容，存在已知重大问题(如休眠不稳定、摄像头/指纹不工作等)"),
        ("★☆☆☆☆ (1星)", "不推荐", "至少1个关键硬件无法在Linux下工作，或无任何可获取的兼容性数据"),
    ]

    rt_headers = ["评分", "级别", "说明"]
    for i, h in enumerate(rt_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.border = THIN_BORDER
    row += 1

    for rating, level, desc in rating_criteria:
        for i, val in enumerate([rating, level, desc], 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.border = THIN_BORDER
            # 评分单元格着色
            if i == 1:
                bg = RATING_COLORS.get(rating, "C0C0C0")
                fg = RATING_FONT_COLORS.get(rating, "000000")
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=fg)
        row += 1

    # 条件格式说明
    row += 2
    ws.cell(row=row, column=1, value="── 条件格式说明 ──").font = Font(
        name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
    row += 1

    legend_text = [
        "• Linux兼容性评分列: 按星级着色（★★★★★=绿, ★★★★☆=浅绿, ★★★☆☆=黄, ★★☆☆☆=橙, ★☆☆☆☆=红, N/A=灰）",
        "• Ubuntu认证列: 是=绿色, 否=红色, 未确认=黄色",
        "• Fedora兼容性列: certified/ships_fedora=绿, friendly=浅绿, community_good=淡绿, community_warning=橙, unknown/N/A=灰",
        "• 数据可信度列: 中高=绿, 中=浅绿, 低=黄",
        "",
        "ⓘ 以上条件格式已自动应用到数据Sheet，无需手动设置。",
        "ⓘ 本文件使用openpyxl生成，无需VBA宏，跨平台兼容 (Windows/macOS/Linux)。",
    ]
    for text in legend_text:
        ws.cell(row=row, column=1, value=text).font = Font(
            name="Microsoft YaHei", size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15


# ── 主流程 ─────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("ThinkPad调研项目 - 最终Excel生成")
    print("=" * 60)

    # 1. 加载数据
    print(f"\n[1/5] 加载合并数据: {INPUT_PATH}")
    with open(INPUT_PATH, "r", encoding="utf-8") as f:
        raw_data = json.load(f)

    models = raw_data["models"]
    meta = raw_data["meta"]
    print(f"  ✓ 加载 {len(models)} 个型号, {len(models[0])} 个字段")

    # 2. 构建DataFrame
    print("[2/5] 构建DataFrame并扁平化数据...")
    df = prepare_dataframe(models)
    print(f"  ✓ DataFrame: {df.shape[0]} 行 × {df.shape[1]} 列")

    # 3. 写入Excel
    print(f"[3/5] 写入Excel到: {OUTPUT_PATH}")
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ThinkPad调研数据")

    # 4. 应用样式和条件格式 (使用openpyxl直接操作)
    print("[4/5] 应用样式和条件格式...")
    from openpyxl import load_workbook
    wb = load_workbook(OUTPUT_PATH)
    ws = wb["ThinkPad调研数据"]

    # 冻结首行
    ws.freeze_panes = "A2"

    # 启用筛选
    ws.auto_filter.ref = ws.dimensions

    # 标题行样式
    num_cols = len(df.columns)
    apply_header_style(ws, num_cols)

    # 数据区域样式
    num_rows = len(df)
    apply_data_style(ws, num_rows, num_cols)

    # 列宽
    auto_width(ws)

    # 条件格式
    apply_conditional_formatting(ws, df)

    # 设置行高
    ws.row_dimensions[1].height = 30

    # 5. 创建数据说明Sheet
    print("[5/5] 创建数据说明Sheet...")
    create_data_description_sheet(wb)

    # 保存
    wb.save(OUTPUT_PATH)
    print(f"\n{'=' * 60}")
    print(f"✅ 完成! 文件已保存: {OUTPUT_PATH}")
    print(f"   Sheet 1: ThinkPad调研数据 ({num_rows}行 × {num_cols}列)")
    print(f"   Sheet 2: 数据说明")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
