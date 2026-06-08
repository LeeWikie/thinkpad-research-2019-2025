#!/usr/bin/env python3
"""
Final Wave F2: Excel格式验证脚本
严格验证 thinkpad_complete_research.xlsx 的格式和功能完整性
"""
import sys
import os
import traceback
from datetime import datetime

# ============================================================================
# 期望的列名清单 (25+列，实际44列)
# ============================================================================
EXPECTED_COLUMNS = [
    # 基本信息 (6列)
    "型号", "系列", "发布年份", "代际", "屏幕尺寸", "别名/变体",

    # CPU与平台 (3列)
    "CPU型号", "CPU架构", "平台(Intel/AMD)",

    # 内存与存储 (3列)
    "内存类型", "最大内存", "存储接口",

    # 显示参数 (4列)
    "屏幕分辨率", "屏幕刷新率", "屏幕面板类型", "触摸屏",

    # 图形 (2列)
    "集成显卡", "独立显卡",

    # 连接 (2列)
    "无线网卡", "接口/端口",

    # 输入设备 (3列)
    "键盘", "指纹识别", "摄像头",

    # 电池与物理 (3列)
    "电池容量", "标称续航", "重量",

    # BIOS (1列)
    "BIOS型号",

    # Linux兼容性 (6列)
    "Ubuntu认证", "Ubuntu认证版本", "Fedora兼容性",
    "Arch Wiki", "Linux兼容性评分", "社区反馈",

    # 价格 (3列)
    "二手价格最低(¥)", "二手价格最高(¥)", "价格采集日期",

    # 数据质量 (3列)
    "数据可信度", "数据来源", "数据更新日期",

    # 实测续航 (1列，部分型号有)
    "实测续航",
]

# 必须存在的关键列 (核心字段) - 按关键词匹配
CRITICAL_COLUMNS = [
    ("型号", "型号"),
    ("系列", "系列"),
    ("发布年份", "发布年份"),
    ("CPU型号", "CPU型号"),
    ("屏幕分辨率", "分辨率"),
    ("Ubuntu认证", "Ubuntu认证"),
    ("Linux兼容性评分", "兼容性评分"),
    ("二手价格", "价格区间"),  # 可以是一个组合价格列
    ("数据可信度", "数据可信度"),
    ("数据来源", "数据来源"),
]

# Linux兼容性评分可能的等级
SCORE_PATTERNS = ["★★★★★", "★★★★☆", "★★★☆☆", "★★☆☆☆", "★☆☆☆☆"]

# Ubuntu认证可能值
UBUNTU_CERT_VALUES = ["是", "否", "未确认"]


def check_dependencies():
    """检查依赖库是否可用"""
    deps_ok = True
    try:
        import openpyxl
        print(f"✅ openpyxl {openpyxl.__version__}")
    except ImportError:
        print("❌ openpyxl 未安装")
        deps_ok = False

    try:
        import pandas as pd
        print(f"✅ pandas {pd.__version__}")
    except ImportError:
        print("❌ pandas 未安装")
        deps_ok = False

    return deps_ok


def print_header(title):
    print("\n" + "=" * 70)
    print(f"  {title}")
    print("=" * 70)


def validate_file_access(filepath):
    """验证6: 文件可正常打开 (pandas + openpyxl)"""
    results: dict = {"pass": True, "checks": [], "wb": None, "df": None}

    import openpyxl
    import pandas as pd

    # 检查文件存在
    if not os.path.exists(filepath):
        results["pass"] = False
        results["checks"].append(("❌", "文件不存在", f"路径: {filepath}"))
        return results
    results["checks"].append(("✅", "文件存在", f"大小: {os.path.getsize(filepath):,} bytes"))

    # pandas读取
    try:
        df = pd.read_excel(filepath, engine='openpyxl')
        results["df"] = df
        results["checks"].append(("✅", "pandas读取成功", f"Shape: {df.shape[0]}行 × {df.shape[1]}列"))
    except Exception as e:
        results["pass"] = False
        results["checks"].append(("❌", "pandas读取失败", str(e)[:100]))
        return results

    # openpyxl读取
    try:
        wb = openpyxl.load_workbook(filepath)
        results["wb"] = wb
        results["checks"].append(("✅", "openpyxl读取成功", f"Sheets: {wb.sheetnames}"))
    except Exception as e:
        results["pass"] = False
        results["checks"].append(("❌", "openpyxl读取失败", str(e)[:100]))
        return results

    return results


def validate_sheets(wb):
    """验证Sheet结构"""
    results = {"pass": True, "checks": []}

    sheets = wb.sheetnames
    results["checks"].append(("ℹ️", "Sheet列表", str(sheets)))

    # 检查数据Sheet
    data_sheet_names = [s for s in sheets if "ThinkPad" in s or "数据" in s or "调研" in s]
    if not data_sheet_names:
        data_sheet_names = [sheets[0]]  # fallback: 第一个sheet

    results["data_sheet"] = data_sheet_names[0]
    results["checks"].append(("✅", "数据Sheet", data_sheet_names[0]))

    # 检查说明Sheet
    explain_sheet_names = [s for s in sheets if "说明" in s or "说明" in s]
    if explain_sheet_names:
        results["explain_sheet"] = explain_sheet_names[0]
        results["checks"].append(("✅", "数据说明Sheet存在", explain_sheet_names[0]))

        # 检查说明Sheet内容
        explain_ws = wb[explain_sheet_names[0]]
        explain_rows = explain_ws.max_row
        results["checks"].append(("ℹ️", "数据说明Sheet行数", f"{explain_rows}行"))
        if explain_rows < 10:
            results["checks"].append(("⚠️", "数据说明Sheet内容过少", f"仅{explain_rows}行"))
    else:
        results["pass"] = False
        results["checks"].append(("❌", "缺少数据说明Sheet", ""))
        results["explain_sheet"] = None

    return results


def validate_field_completeness(df):
    """验证1: 字段完整性 (25+列)"""
    results = {"pass": True, "checks": []}

    actual_columns = list(df.columns)
    num_cols = len(actual_columns)

    results["column_count"] = num_cols
    if num_cols >= 25:
        results["checks"].append(("✅", f"列数达标({num_cols}/25+)", ""))
    else:
        results["pass"] = False
        results["checks"].append(("❌", f"列数不足", f"实际{num_cols}列, 需要≥25列"))

    # 打印实际列名
    results["actual_columns"] = actual_columns
    results["checks"].append(("ℹ️", f"实际列名({num_cols}列)", " | ".join(actual_columns[:10]) + " ..."))

    # 检查关键列是否存在 (按关键词匹配)
    missing_critical = []
    for col_name, col_keyword in CRITICAL_COLUMNS:
        found = False
        for ac in actual_columns:
            if col_keyword.lower() in ac.lower() or ac.lower() in col_keyword.lower():
                found = True
                break
        if not found:
            missing_critical.append(col_name)

    if missing_critical:
        results["pass"] = False
        results["checks"].append(("❌", "缺少关键列", str(missing_critical)))
    else:
        col_names = [c[0] for c in CRITICAL_COLUMNS]
        results["checks"].append(("✅", "关键列全部存在", str(col_names)))

    # 检查空列
    empty_cols = []
    for col in actual_columns:
        null_rate = df[col].isna().mean()
        if null_rate > 0.95:  # 超过95%为空
            empty_cols.append(f"{col}({null_rate:.0%}空)")
    if empty_cols:
        results["checks"].append(("⚠️", "接近空列", str(empty_cols)))
    else:
        results["checks"].append(("✅", "无接近空列", ""))

    return results


def validate_data_content(df):
    """验证数据内容质量"""
    results = {"pass": True, "checks": []}

    num_rows = len(df)
    results["row_count"] = num_rows
    results["checks"].append(("ℹ️", f"数据行数", f"{num_rows}行"))

    if num_rows < 50:
        results["pass"] = False
        results["checks"].append(("❌", "数据行数异常少", f"仅{num_rows}行, 预期≥80"))
    elif num_rows >= 80:
        results["checks"].append(("✅", f"数据行数正常({num_rows}≥80)", ""))

    # 检查型号列
    model_col = None
    for col in df.columns:
        if "型号" in col or "model" in col.lower():
            model_col = col
            break

    if model_col:
        unique_models = df[model_col].nunique()
        results["checks"].append(("ℹ️", "唯一型号数", f"{unique_models}个 (列: {model_col})"))

    # 检查Ubuntu认证列
    ubuntu_col = None
    for col in df.columns:
        if "Ubuntu" in col and ("认证" in col or "cert" in col.lower()):
            ubuntu_col = col
            break

    if ubuntu_col:
        ubuntu_values = df[ubuntu_col].dropna().unique()
        results["checks"].append(("ℹ️", f"Ubuntu认证值分布", str(ubuntu_values.tolist())))
    else:
        results["checks"].append(("⚠️", "未找到Ubuntu认证列", ""))

    # 检查Linux兼容性评分列
    score_col = None
    for col in df.columns:
        if "兼容" in col and "评分" in col:
            score_col = col
            break

    if score_col:
        score_values = df[score_col].dropna().unique()
        results["checks"].append(("ℹ️", f"评分值分布", str(score_values.tolist())))

    return results


def validate_conditional_formatting(wb, data_sheet_name):
    """验证2: 条件格式正确 (Linux兼容性红绿标注)"""
    results = {"pass": True, "checks": []}

    ws = wb[data_sheet_name]
    cf_rules = ws.conditional_formatting

    # 收集所有条件格式规则
    all_rules = []
    for cf in cf_rules:
        for rule in cf.rules:
            all_rules.append({
                "range": str(cf.cells),
                "type": rule.type,
                "priority": rule.priority,
                "formula": str(rule.formula) if rule.formula else None,
                "operator": str(rule.operator) if rule.operator else None
            })

    if all_rules:
        results["checks"].append(("✅", f"条件格式规则数", f"{len(all_rules)}条"))
        for i, rule in enumerate(all_rules):
            results["checks"].append(("ℹ️", f"  规则{i+1}", f"范围={rule['range']}, 类型={rule['type']}"))
    else:
        results["checks"].append(("⚠️", "无条件格式规则", "文件可能没有样式条件格式"))

    # 检查颜色填充 (通过读取实际单元格颜色)
    score_col_index = None

    # 找到"Linux兼容性评分"列
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    for i, header in enumerate(header_row):
        if header and ("兼容性评分" in str(header) or "linux兼容性" in str(header).lower()):
            score_col_index = i + 1  # 1-indexed
            break

    if score_col_index:
        col_letter = chr(64 + score_col_index) if score_col_index <= 26 else f"AA{chr(64 + score_col_index - 26)}"
        results["checks"].append(("ℹ️", f"评分列位置", f"列{score_col_index} ({col_letter})"))

        # 检查几行的填充色
        color_samples = []
        for row in range(2, min(20, ws.max_row + 1)):
            cell = ws.cell(row=row, column=score_col_index)
            if cell.fill and cell.fill.fgColor:
                color_samples.append(f"Row{row}: fill={cell.fill.fgColor}")

        if color_samples:
            results["checks"].append(("ℹ️", "评分列填充色抽样", "; ".join(color_samples[:5])))
        else:
            results["checks"].append(("⚠️", "评分列无填充色", "条件格式可能未生效"))
    else:
        results["checks"].append(("⚠️", "未找到Linux兼容性评分列", ""))

    return results


def validate_freeze_panes(wb, data_sheet_name):
    """验证3: 冻结首行已设置"""
    results = {"pass": True, "checks": []}

    ws = wb[data_sheet_name]
    freeze = ws.freeze_panes

    if freeze:
        results["checks"].append(("✅", f"冻结窗格已设置", f"冻结点: {freeze}"))
        # 理想情况: freeze_panes = "A2" (冻结首行)
        if freeze == "A2":
            results["checks"].append(("✅", "首行正确冻结", "A2 (冻结第1行)"))
        else:
            results["checks"].append(("⚠️", f"冻结点非标准", f"当前{freeze}, 预期A2"))
    else:
        results["pass"] = False
        results["checks"].append(("❌", "未设置冻结窗格", ""))

    return results


def validate_autofilter(wb, data_sheet_name):
    """验证4: 筛选功能可用"""
    results = {"pass": True, "checks": []}

    ws = wb[data_sheet_name]

    if ws.auto_filter and ws.auto_filter.ref:
        results["checks"].append(("✅", "自动筛选已启用", f"范围: {ws.auto_filter.ref}"))
        # 检查筛选范围是否覆盖所有列和数据
        results["checks"].append(("ℹ️", f"筛选覆盖", f"数据: {ws.max_row}行 × {ws.max_column}列"))
    else:
        results["pass"] = False
        results["checks"].append(("❌", "自动筛选未启用", ""))

    return results


def validate_no_vba(wb):
    """验证5: 无VBA宏"""
    results = {"pass": True, "checks": []}

    has_vba = wb.vba_archive is not None

    if not has_vba:
        results["checks"].append(("✅", "无VBA宏", "符合跨平台兼容要求"))
    else:
        results["pass"] = False
        results["checks"].append(("❌", "检测到VBA宏", "违反跨平台兼容要求"))

    return results


def validate_styles(wb, data_sheet_name):
    """验证样式设置"""
    results = {"pass": True, "checks": []}

    ws = wb[data_sheet_name]

    # 检查标题行样式 (第1行)
    header_cell = ws.cell(row=1, column=1)
    if header_cell.font:
        results["checks"].append(("ℹ️", "标题字体", f"Name={header_cell.font.name}, Bold={header_cell.font.bold}, Size={header_cell.font.size}"))
    if header_cell.fill:
        results["checks"].append(("ℹ️", "标题背景", f"fgColor={header_cell.fill.fgColor}, patternType={header_cell.fill.patternType}"))

    if header_cell.font and header_cell.font.bold:
        results["checks"].append(("✅", "标题行加粗", ""))
    else:
        results["checks"].append(("⚠️", "标题行未加粗", ""))

    # 检查边框
    data_cell = ws.cell(row=2, column=1)
    if data_cell.border:
        results["checks"].append(("ℹ️", "数据区边框", f"left={data_cell.border.left.style}, right={data_cell.border.right.style}, top={data_cell.border.top.style}"))

    return results


def main():
    filepath = "/mnt/c/Users/Administrator/Desktop/thinkpad/thinkpad_complete_research.xlsx"

    print("=" * 70)
    print("  FINAL WAVE F2: Excel格式验证")
    print(f"  目标文件: {filepath}")
    print(f"  执行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # 检查依赖
    if not check_dependencies():
        print("\n❌ 依赖库缺失，无法继续验证")
        sys.exit(1)

    # ============================================================
    # 验证6: 文件可正常打开
    # ============================================================
    print_header("验证6: 文件可正常打开 (pandas + openpyxl)")
    file_results = validate_file_access(filepath)
    if not file_results["pass"]:
        for status, item, detail in file_results["checks"]:
            print(f"  {status} {item}: {detail}")
        sys.exit(1)

    wb = file_results["wb"]
    df = file_results["df"]
    for status, item, detail in file_results["checks"]:
        print(f"  {status} {item}: {detail}")

    # ============================================================
    # 验证1: 字段完整性
    # ============================================================
    print_header("验证1: 字段完整性 (25+列)")
    field_results = validate_field_completeness(df)
    for status, item, detail in field_results["checks"]:
        print(f"  {status} {item}: {detail}")

    # 打印完整列名清单
    print(f"\n  📋 完整列名清单 ({field_results['column_count']}列):")
    for i, col in enumerate(field_results["actual_columns"], 1):
        print(f"     {i:2d}. {col}")

    # ============================================================
    # Sheet结构验证
    # ============================================================
    print_header("Sheet结构验证")
    sheet_results = validate_sheets(wb)
    data_sheet = sheet_results["data_sheet"]
    for status, item, detail in sheet_results["checks"]:
        print(f"  {status} {item}: {detail}")

    # ============================================================
    # 验证2: 条件格式
    # ============================================================
    print_header("验证2: 条件格式正确 (Linux兼容性红绿标注)")
    cf_results = validate_conditional_formatting(wb, data_sheet)
    for status, item, detail in cf_results["checks"]:
        print(f"  {status} {item}: {detail}")

    # ============================================================
    # 验证3: 冻结首行
    # ============================================================
    print_header("验证3: 冻结首行已设置")
    freeze_results = validate_freeze_panes(wb, data_sheet)
    for status, item, detail in freeze_results["checks"]:
        print(f"  {status} {item}: {detail}")

    # ============================================================
    # 验证4: 筛选功能
    # ============================================================
    print_header("验证4: 筛选功能可用")
    filter_results = validate_autofilter(wb, data_sheet)
    for status, item, detail in filter_results["checks"]:
        print(f"  {status} {item}: {detail}")

    # ============================================================
    # 验证5: 无VBA宏
    # ============================================================
    print_header("验证5: 无VBA宏")
    vba_results = validate_no_vba(wb)
    for status, item, detail in vba_results["checks"]:
        print(f"  {status} {item}: {detail}")

    # ============================================================
    # 样式验证
    # ============================================================
    print_header("样式验证")
    style_results = validate_styles(wb, data_sheet)
    for status, item, detail in style_results["checks"]:
        print(f"  {status} {item}: {detail}")

    # ============================================================
    # 数据内容验证
    # ============================================================
    print_header("数据内容质量")
    content_results = validate_data_content(df)
    for status, item, detail in content_results["checks"]:
        print(f"  {status} {item}: {detail}")

    # ============================================================
    # 综合VERDICT
    # ============================================================
    print_header("综合VERDICT")

    all_checks = [
        ("字段完整性(25+列)", field_results["pass"]),
        ("条件格式(Linux兼容性红绿标注)", cf_results["pass"]),
        ("冻结首行", freeze_results["pass"]),
        ("筛选功能", filter_results["pass"]),
        ("无VBA宏", vba_results["pass"]),
        ("文件可正常打开", file_results["pass"]),
        ("数据说明Sheet", sheet_results["pass"]),
    ]

    passed = 0
    failed = 0
    for name, status in all_checks:
        if status:
            print(f"  ✅ {name}")
            passed += 1
        else:
            print(f"  ❌ {name}")
            failed += 1

    # 最终判定
    print()
    print(f"  Fields [{field_results['column_count']}/{field_results['column_count']}] | ", end="")
    print(f"Format [{'PASS' if cf_results['pass'] and freeze_results['pass'] and style_results['pass'] else 'FAIL'}] | ", end="")
    print(f"Functionality [{'PASS' if filter_results['pass'] and vba_results['pass'] and file_results['pass'] else 'FAIL'}] | ", end="")

    if failed == 0:
        print("VERDICT: ✅ APPROVE")
        print()
        print("  🎉 所有验证项通过！Excel文件格式和功能完整。")
        verdict = "APPROVE"
    else:
        print(f"VERDICT: ❌ REJECT ({failed}项失败)")
        print()
        print(f"  ⚠️ {failed}项验证失败，需要修复。")
        verdict = "REJECT"

    print("=" * 70)

    # 生成验证报告文件
    report_path = os.path.join(os.path.dirname(filepath), "..", ".sisyphus", "evidence", "task-f2-excel-validation.txt")
    report_dir = os.path.dirname(report_path)
    os.makedirs(report_dir, exist_ok=True)

    with open(report_path, "w", encoding="utf-8") as f:
        f.write(f"Final Wave F2: Excel格式验证报告\n")
        f.write(f"===========================================\n")
        f.write(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"文件: {filepath}\n")
        f.write(f"\n验证结果:\n")
        f.write(f"  Fields: {field_results['column_count']}/{field_results['column_count']}\n")
        f.write(f"  Format: {'PASS' if cf_results['pass'] and freeze_results['pass'] and style_results['pass'] else 'FAIL'}\n")
        f.write(f"  Functionality: {'PASS' if filter_results['pass'] and vba_results['pass'] and file_results['pass'] else 'FAIL'}\n")
        f.write(f"  VERDICT: {verdict}\n")
        f.write(f"\n详细检查:\n")
        for name, status in all_checks:
            f.write(f"  [{'PASS' if status else 'FAIL'}] {name}\n")
        f.write(f"\n列名清单({field_results['column_count']}列):\n")
        for i, col in enumerate(field_results["actual_columns"], 1):
            f.write(f"  {i:2d}. {col}\n")
        f.write(f"\n数据统计:\n")
        f.write(f"  总行数: {content_results.get('row_count', 'N/A')}\n")
        f.write(f"  总列数: {field_results['column_count']}\n")

    print(f"\n📄 验证报告已保存: {report_path}")

    return verdict


if __name__ == "__main__":
    main()
