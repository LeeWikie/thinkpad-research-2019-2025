#!/usr/bin/env python3
"""F1: 数据完整性审计 - Oracle Agent"""

import json, glob, sys, os
from collections import Counter, defaultdict
from datetime import datetime

os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# ============ 1. Load Model Lists ============
def load_model_lists():
    """Load all model list JSON files and return unified set of expected models."""
    all_models = {}  # model_name -> {series, year, generation}
    series_counts = Counter()
    year_counts = Counter()

    for fpath in glob.glob("data/raw/models_*.json"):
        with open(fpath) as f:
            data = json.load(f)

        models_key = None
        if isinstance(data, list):
            models = data
        elif "models" in data:
            models = data["models"]
        else:
            print(f"  WARN: Unknown format: {fpath}")
            continue

        for m in models:
            name = m["model"].replace("ThinkPad ", "")
            all_models[name] = {
                "series": m.get("series", "?"),
                "year": m.get("year", 0),
                "generation": m.get("generation", ""),
                "source": os.path.basename(fpath)
            }
            series_counts[m.get("series", "?")] += 1
            year_counts[str(m.get("year", "?"))] += 1

    return all_models, series_counts, year_counts

expected_models, expected_series, expected_years = load_model_lists()
print(f"=== MODEL COVERAGE ===")
print(f"  Expected models from JSON: {len(expected_models)}")
print(f"  By series: {dict(expected_series)}")
print(f"  By year:   {dict(sorted(expected_years.items()))}")
print()

# ============ 2. Load Raw Data for Cross-Check ============
def load_raw_specs():
    """Load all spec JSON files for cross-referencing."""
    raw_data = {}
    for fpath in glob.glob("data/raw/specs_*.json"):
        with open(fpath) as f:
            data = json.load(f)
        for m in data.get("models", []):
            name = m["model"].replace("ThinkPad ", "")
            raw_data[name] = m
    return raw_data

def load_raw_compat():
    """Load compatibility data."""
    compat_data = {}
    for fname in ["ubuntu_cert.json", "fedora_compat.json", "arch_compat.json", "community_compat.json", "prices_cn.json"]:
        fpath = f"data/raw/{fname}"
        if not os.path.exists(fpath):
            continue
        with open(fpath) as f:
            data = json.load(f)
        compat_data[fname] = data
    return compat_data

raw_specs = load_raw_specs()
raw_compat = load_raw_compat()
print(f"  Raw spec models loaded: {len(raw_specs)}")
print(f"  Raw compat files: {list(raw_compat.keys())}")
print()

# ============ 3. Load Excel ============
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

wb = openpyxl.load_workbook("thinkpad_complete_research.xlsx")
ws = wb["ThinkPad调研数据"]

# Read headers
headers = [cell.value for cell in ws[1]]
print(f"=== EXCEL STRUCTURE ===")
print(f"  Sheet: {ws.title}")
print(f"  Dimensions: {ws.dimensions}")
print(f"  Headers: {len(headers)} columns")
print(f"  Column names: {headers}")
print()

# Read data rows
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    rows.append(dict(zip(headers, row)))

print(f"  Data rows: {len(rows)}")
print()

# ============ 4. Must-Have Verification ============
print("=" * 70)
print("MUST-HAVE CHECKLIST VERIFICATION")
print("=" * 70)

# 4a. 系列覆盖
excel_series = Counter()
for r in rows:
    excel_series[r.get("系列", "?")] += 1

required_series = ["T", "X", "P", "E", "L", "Z", "X1"]
missing_series = []
for s in required_series:
    if s not in excel_series:
        missing_series.append(s)
    elif s == "X1":
        # Check X1 sub-variants
        x1_models = [r for r in rows if "X1" in str(r.get("系列", ""))]
        if not x1_models:
            missing_series.append("X1 (sub-variants)")

print(f"\n[1/7] 全系列型号覆盖 (T/X/P/E/L/Z/X1)")
print(f"  要求: T/X/P/E/L/Z/X1 系列，2019-2025")
print(f"  Excel系列分布: {dict(excel_series)}")
print(f"  缺失系列: {missing_series if missing_series else 'NONE'}")
print(f"  年份范围: {min(r.get('发布年份', 0) for r in rows)}-{max(r.get('发布年份', 0) for r in rows)}")

# 4b. 型号覆盖率
excel_model_names = set()
for r in rows:
    mn = r.get("型号", "")
    excel_model_names.add(mn)

expected_names = set(expected_models.keys())
missing_in_excel = expected_names - excel_model_names
extra_in_excel = excel_model_names - expected_names

print(f"\n  型号覆盖率:")
print(f"  预期: {len(expected_names)} | Excel: {len(excel_model_names)}")
if missing_in_excel:
    print(f"  缺失型号 ({len(missing_in_excel)}): {sorted(missing_in_excel)}")
if extra_in_excel:
    print(f"  额外型号 ({len(extra_in_excel)}): {sorted(extra_in_excel)}")

# 4c. Official Hardware Parameters Completeness
print(f"\n[2/7] 官方硬件参数100%完整")
critical_fields = {
    "CPU型号": "cpu",
    "分辨率": "resolution",
    "尺寸(英寸)": "screen_size",
    "内存类型": "memory",
    "最大内存": "max_memory",
    "存储接口": "storage",
    "接口": "ports",
    "指纹识别": "fingerprint",
    "摄像头": "webcam",
    "集成显卡": "igpu",
    "独立显卡": "dgpu",
    "无线": "wireless",
}

completeness = {}
for field in critical_fields:
    missing_count = sum(1 for r in rows if not r.get(field) or r.get(field) in ("N/A", "N/A", None, ""))
    total = len(rows)
    pct = (total - missing_count) / total * 100
    completeness[field] = {"present": total - missing_count, "total": total, "pct": round(pct, 1)}
    if missing_count > 0:
        print(f"  ❌ {field}: {total - missing_count}/{total} ({pct:.1f}%) - {missing_count} missing")
        if missing_count <= 5:
            for r in rows:
                if not r.get(field) or r.get(field) in ("N/A", "", None):
                    print(f"     - {r.get('型号', '?')}")

# Also check refresh rate, touch screen, official battery life, weight, BIOS
extra_fields = ["刷新率(Hz)", "触摸屏", "官方标称续航(h)", "重量"]
for field in extra_fields:
    missing_count = sum(1 for r in rows if not r.get(field) or r.get(field) in ("N/A", "", None))
    total = len(rows)
    pct = (total - missing_count) / total * 100
    completeness[field] = {"present": total - missing_count, "total": total, "pct": round(pct, 1)}
    if missing_count > total * 0.2:
        print(f"  ⚠️  {field}: {total - missing_count}/{total} ({pct:.1f}%) - {missing_count} missing")

# 4d. Ubuntu Certified
print(f"\n[3/7] Ubuntu Certified认证状态")
ubuntu_status = Counter()
for r in rows:
    ubuntu_status[str(r.get("Ubuntu认证", "?"))] += 1
print(f"  分布: {dict(ubuntu_status)}")
ubuntu_yes = ubuntu_status.get("是", 0)
ubuntu_no = ubuntu_status.get("否", 0)
ubuntu_na = ubuntu_status.get("N/A", 0) + ubuntu_status.get("未确认", 0)
print(f"  是={ubuntu_yes} | 否={ubuntu_no} | 未确认/NA={ubuntu_na}")

# Check against raw ubuntu cert data
if "ubuntu_cert.json" in raw_compat:
    uc = raw_compat["ubuntu_cert.json"]
    cert_models = set()
    for m in uc.get("models", uc if isinstance(uc, list) else []):
        cert_models.add(m.get("model", "").replace("ThinkPad ", ""))
    print(f"  原始认证模型数: {len(cert_models)}")

# 4e. Linux Compatibility Score
print(f"\n[4/7] Linux兼容性综合评分")
linux_scores = Counter()
for r in rows:
    linux_scores[str(r.get("Linux兼容性评分", "?"))] += 1
print(f"  分布: {dict(linux_scores)}")

# Check against raw community + fedora + arch data
community_data = raw_compat.get("community_compat.json", {})
fedora_data = raw_compat.get("fedora_compat.json", {})
arch_data = raw_compat.get("arch_compat.json", {})
print(f"  Community compat raw entries: {len(community_data.get('models', []) if isinstance(community_data, dict) else community_data)}")
print(f"  Fedora raw entries: {len(fedora_data.get('models', []) if isinstance(fedora_data, dict) else fedora_data)}")
print(f"  Arch raw entries: {len(arch_data.get('models', []) if isinstance(arch_data, dict) else arch_data)}")

# 4f. Chinese Used Price
print(f"\n[5/7] 中国二手价格区间")
price_present = sum(1 for r in rows if r.get("二手价格区间(CNY)") and r.get("二手价格区间(CNY)") not in ("N/A", "", None))
price_na = sum(1 for r in rows if not r.get("二手价格区间(CNY)") or r.get("二手价格区间(CNY)") in ("N/A", "", None))
print(f"  有价格: {price_present}/{len(rows)} | N/A: {price_na}/{len(rows)}")

# Sample prices
prices = [r.get("二手价格区间(CNY)") for r in rows if r.get("二手价格区间(CNY)") and r.get("二手价格区间(CNY)") not in ("N/A", "")]
print(f"  样例: {prices[:5]}")

# 4g. Excel Conditional Formatting
print(f"\n[6/7] Excel条件格式（Linux兼容性红绿标注）")
if ws.conditional_formatting:
    cf_rules = []
    for cf in ws.conditional_formatting:
        cf_rules.append(str(cf))
    print(f"  条件格式规则数: {len(ws.conditional_formatting)}")
    print(f"  覆盖范围: {[str(cf) for cf in ws.conditional_formatting][:3]}...")
else:
    print(f"  ❌ 无条件格式!")

# Also check freeze panes
print(f"  冻结窗格: {ws.freeze_panes}")

# 4h. Data Source Traceability
print(f"\n[7/7] 数据来源可追溯")
source_present = sum(1 for r in rows if r.get("数据来源URL") and r.get("数据来源URL") not in ("N/A", "", None))
source_na = sum(1 for r in rows if not r.get("数据来源URL") or r.get("数据来源URL") in ("N/A", "", None))
print(f"  有来源URL: {source_present}/{len(rows)} | N/A: {source_na}/{len(rows)}")

# Check data update date
date_present = sum(1 for r in rows if r.get("数据更新日期") and r.get("数据更新日期") not in ("N/A", "", None))
print(f"  有数据更新日期: {date_present}/{len(rows)}")

# ============ 5. N/A Ratio Analysis ============
print(f"\n{'=' * 70}")
print("N/A RATIO ANALYSIS (按字段)")
print("=" * 70)

total_rows = len(rows)
na_stats = {}
for h in headers:
    na_count = sum(1 for r in rows if not r.get(h) or str(r.get(h)).strip() in ("N/A", "", "无", "None"))
    na_pct = na_count / total_rows * 100
    na_stats[h] = {"count": na_count, "pct": round(na_pct, 1)}

# Show high N/A fields
print(f"\n  ALL FIELDS N/A RATIOS:")
print(f"  {'Field':<30} {'N/A':>5} {'%':>6} {'Status'}")
print(f"  {'-'*50}")
high_na_fields = []
for h, s in sorted(na_stats.items(), key=lambda x: -x[1]["pct"]):
    status = "🔴 HIGH" if s["pct"] > 50 else ("🟡 MED" if s["pct"] > 20 else "🟢 OK")
    if s["pct"] > 20:
        high_na_fields.append((h, s))
    print(f"  {h:<30} {s['count']:>5} {s['pct']:>5.1f}% {status}")

overall_na = sum(s["count"] for s in na_stats.values())
overall_cells = total_rows * len(headers)
overall_na_pct = overall_na / overall_cells * 100
print(f"\n  OVERALL N/A: {overall_na}/{overall_cells} = {overall_na_pct:.1f}%")

# ============ 6. FABRICATION DETECTION ============
print(f"\n{'=' * 70}")
print("FABRICATION DETECTION (编造数据检查)")
print("=" * 70)

fabrication_issues = []
sample_size = min(30, len(rows))
checked_fields = ["CPU型号", "分辨率", "尺寸(英寸)", "Ubuntu认证", "二手价格区间(CNY)"]

# Cross-check CPU models against raw specs
import re
cpu_discrepancies = []
for r in rows:
    model = r.get("型号", "")
    if not model:
        continue
    # Find in raw specs
    raw = raw_specs.get(model, None)
    if raw:
        raw_cpu = raw.get("cpu", {})
        if isinstance(raw_cpu, dict):
            raw_cpu_str = raw_cpu.get("family", "") + " " + ", ".join(raw_cpu.get("models", []))
        else:
            raw_cpu_str = str(raw_cpu)
        excel_cpu = str(r.get("CPU型号", ""))

        if excel_cpu in ("N/A", "", "None", None) and raw_cpu_str.strip():
            cpu_discrepancies.append(f"  ❌ {model}: Excel=N/A but raw data has CPU: {raw_cpu_str[:80]}")
        elif raw_cpu_str.strip() and excel_cpu not in ("N/A", "", "None") and not any(m in excel_cpu for m in ["i5-", "i7-", "i9-", "Ryzen", "Intel"]) :
            pass  # fine
    else:
        # Model not in raw specs - check if it exists in model lists
        if model in expected_models:
            cpu_discrepancies.append(f"  ⚠️  {model}: No raw spec data, but exists in model list")

for disc in cpu_discrepancies[:20]:
    print(disc)
if len(cpu_discrepancies) > 20:
    print(f"  ... and {len(cpu_discrepancies) - 20} more")

# Check Ubuntu certification has corresponding raw data
ubuntu_fab = []
ubuntu_cert_raw = raw_compat.get("ubuntu_cert.json", {})
if ubuntu_cert_raw:
    raw_cert_models = set()
    for m in (ubuntu_cert_raw.get("models", []) if isinstance(ubuntu_cert_raw, dict) else ubuntu_cert_raw):
        raw_cert_models.add(m.get("model", ""))
    
    for r in rows:
        excel_cert = r.get("Ubuntu认证", "N/A")
        model = r.get("型号", "")
        if excel_cert == "是" and model not in raw_cert_models:
            ubuntu_fab.append(f"  ⚠️  {model}: Excel says '是' but not in raw ubuntu_cert.json")

for f in ubuntu_fab[:10]:
    print(f)

# Check prices have CNY format
price_format_issues = []
for r in rows:
    price = str(r.get("二手价格区间(CNY)", ""))
    if price and price != "N/A":
        if "¥" not in price and "CNY" not in price and not re.search(r'\d+', price):
            price_format_issues.append(f"  ⚠️  {r.get('型号')}: Suspicious price format: {price}")

for f in price_format_issues[:10]:
    print(f)

# Check CPU format consistency
cpu_format_issues = []
for r in rows:
    cpu = str(r.get("CPU型号", ""))
    if cpu and cpu not in ("N/A", "None", ""):
        if not re.search(r'(i\d|Ryzen|Core|Intel|AMD|Snapdragon|Apple)', cpu, re.I):
            cpu_format_issues.append(f"  ⚠️  {r.get('型号')}: Unusual CPU format: {cpu}")

for f in cpu_format_issues[:10]:
    print(f)

# ============ 7. RESOLUTION FIELD DEEP CHECK ============
print(f"\n{'=' * 70}")
print("RESOLUTION & CRITICAL FIELD DEEP DIVE")
print("=" * 70)

# Find all models with missing CPU or resolution
critical_missing = []
for r in rows:
    model = r.get("型号", "")
    cpu = str(r.get("CPU型号", ""))
    res = str(r.get("分辨率", ""))
    issues = []
    if cpu in ("N/A", "", "None"):
        issues.append("CPU missing")
    if res in ("N/A", "", "None"):
        issues.append("Resolution missing")
    if issues:
        critical_missing.append((model, issues, r.get("系列", ""), r.get("发布年份", "")))

print(f"Models with CRITICAL missing data: {len(critical_missing)}")
for model, issues, series, year in critical_missing:
    print(f"  ❌ {model} | {series} | {year} | Issues: {', '.join(issues)}")

# ============ 8. VERDICT ============
print(f"\n{'=' * 70}")
print("FINAL VERDICT")
print("=" * 70)

# Score calculation
must_have_scores = {
    "全系列型号覆盖": 1.0 if not missing_in_excel and not missing_series else 0.5,
    "官方硬件参数100%完整": 0.0 if len(critical_missing) > 0 else (1.0 if overall_na_pct < 10 else 0.5),
    "Ubuntu Certified认证状态": 1.0 if ubuntu_na < len(rows) * 0.5 else 0.5,
    "Linux兼容性综合评分": 1.0 if len(linux_scores) > 1 else 0.5,
    "中国二手价格区间": min(1.0, price_present / len(rows)),
    "Excel条件格式": 1.0 if ws.conditional_formatting else 0.0,
    "数据来源可追溯": min(1.0, source_present / len(rows)),
}

print(f"\n  Must-Have Scores:")
total_score = 0
for k, v in must_have_scores.items():
    total_score += v
    print(f"  {'✅' if v >= 0.8 else '❌'} {k}: {v:.1%}")

avg_score = total_score / len(must_have_scores)
print(f"\n  Average Score: {avg_score:.1%}")

has_fabrication = len(cpu_discrepancies) > 0 and len(critical_missing) > 0  # Missing data but raw has it
fabrication_count = len(cpu_discrepancies)

# Final verdict
if avg_score >= 0.8 and not has_fabrication:
    verdict = "APPROVE"
elif avg_score >= 0.6 and len(critical_missing) <= 10:
    verdict = "APPROVE WITH CAVEATS"
else:
    verdict = "REJECT"

# Summary line
na_ratio = f"{overall_na_pct:.1f}%"
fab_status = f"CLEAN" if fabrication_count == 0 else f"{fabrication_count} issues"
print(f"\n  OUTPUT: Must Have [{int(sum(v >= 0.8 for v in must_have_scores.values()))}/{len(must_have_scores)}] | "
      f"N/A Ratio [{na_ratio}] | Fabrication [{fab_status}] | VERDICT: {verdict}")

# ============ 9. Save Report ============
report_path = ".sisyphus/evidence/task-22-audit.txt"
with open(report_path, "w") as f:
    f.write(f"Task 22: 数据完整性审计报告\n")
    f.write(f"审计时间: {datetime.now().isoformat()}\n")
    f.write(f"{'='*70}\n\n")
    f.write(f"VERDICT: {verdict}\n")
    f.write(f"Must Have: {sum(v >= 0.8 for v in must_have_scores.values())}/{len(must_have_scores)}\n")
    f.write(f"N/A Ratio: {na_ratio}\n")
    f.write(f"Fabrication: {fab_status}\n\n")
    f.write(f"Critical Missing Fields: {len(critical_missing)} models\n")
    for model, issues, series, year in critical_missing:
        f.write(f"  - {model} ({series}, {year}): {', '.join(issues)}\n")

print(f"\n  Report saved to: {report_path}")
