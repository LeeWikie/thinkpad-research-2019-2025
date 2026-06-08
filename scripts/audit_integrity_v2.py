#!/usr/bin/env python3
"""F1: 数据完整性审计 - Oracle Agent (v2 - 修复名称匹配)"""

import json, glob, sys, os, re
from collections import Counter, defaultdict
from datetime import datetime

os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def strip_thinkpad(name):
    """Normalize ThinkPad model names for comparison."""
    s = name.strip()
    s = s.replace("ThinkPad ", "").replace("Thinkpad ", "").strip()
    return s

# ============ 1. Load Model Lists ============
def load_model_lists():
    all_models = {}
    series_counts = Counter()
    year_counts = Counter()

    for fpath in glob.glob("data/raw/models_*.json"):
        with open(fpath) as f:
            data = json.load(f)
        models = data if isinstance(data, list) else data.get("models", [])
        for m in models:
            name = strip_thinkpad(m["model"])
            all_models[name] = {
                "series": m.get("series", "?"),
                "year": m.get("year", 0),
                "generation": m.get("generation", ""),
                "source": os.path.basename(fpath)
            }
            series_counts[m.get("series", "?")] += 1
            year_counts[str(m.get("year", "?"))] += 1

    return all_models, series_counts, year_counts

expected_models, expected_series, expected_years = load_model_lists()
print(f"=== MODEL COVERAGE ===")
print(f"  Expected models from JSON: {len(expected_models)}")
print(f"  By series: {dict(expected_series)}")
print(f"  By year:   {dict(sorted(expected_years.items()))}")
print()

# ============ 2. Load Raw Data ============
def load_raw_specs():
    raw_data = {}
    for fpath in glob.glob("data/raw/specs_*.json"):
        with open(fpath) as f:
            data = json.load(f)
        for m in data.get("models", []):
            name = strip_thinkpad(m["model"])
            raw_data[name] = m
    return raw_data

def load_raw_compat():
    compat_data = {}
    for fname in ["ubuntu_cert.json", "fedora_compat.json", "arch_compat.json", "community_compat.json", "prices_cn.json"]:
        fpath = f"data/raw/{fname}"
        if not os.path.exists(fpath):
            continue
        with open(fpath) as f:
            data = json.load(f)
        compat_data[fname] = data
    return compat_data

raw_specs = load_raw_specs()
raw_compat = load_raw_compat()
print(f"  Raw spec models loaded: {len(raw_specs)}")
print(f"  Raw compat files: {list(raw_compat.keys())}")

# ============ 3. Load Excel ============
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)

wb = openpyxl.load_workbook("thinkpad_complete_research.xlsx")
ws = wb["ThinkPad调研数据"]

headers = [cell.value for cell in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    rows.append(dict(zip(headers, row)))

print(f"\n=== EXCEL STRUCTURE ===")
print(f"  Sheet: {ws.title} | {len(rows)} rows x {len(headers)} cols")
print(f"  Freeze: {ws.freeze_panes} | AutoFilter: {ws.auto_filter.ref if ws.auto_filter else 'NONE'}")

# ============ 4. Model Coverage (Fixed) ============
excel_model_names = set()
excel_model_normalized = {}
for r in rows:
    mn = r.get("型号", "")
    excel_model_names.add(mn)
    excel_model_normalized[strip_thinkpad(mn)] = mn

expected_names = set(expected_models.keys())
excel_normalized = set(excel_model_normalized.keys())

missing_in_excel = expected_names - excel_normalized
extra_in_excel = excel_normalized - expected_names

# Categorize missing
missing_by_series = defaultdict(list)
for m in sorted(missing_in_excel):
    info = expected_models.get(m, {})
    missing_by_series[info.get("series", "?")].append(m)

print(f"\n{'=' * 70}")
print("MUST-HAVE CHECKLIST VERIFICATION")
print("=" * 70)

print(f"\n[1/7] 全系列型号覆盖 (T/X/P/E/L/Z/X1, 2019-2025)")
print(f"  预期型号: {len(expected_names)} | Excel型号: {len(excel_normalized)}")
print(f"  缺失: {len(missing_in_excel)} | 额外: {len(extra_in_excel)}")

# Also check by macro-series  
model_series_map = {"T": [], "X": [], "P": [], "E": [], "L": [], "Z": [], "X1": []}
for r in rows:
    s = str(r.get("系列", ""))
    m = r.get("型号", "")
    # Map series names
    if "T Series" in s or s.startswith("T"):
        model_series_map["T"].append(m)
    elif s.startswith("P") or "P Series" in s:
        model_series_map["P"].append(m)
    elif any(s.startswith(x) for x in ["X1 ", "X1"]):
        model_series_map["X1"].append(m)
    elif s.startswith("X"):
        model_series_map["X"].append(m)
    elif s.startswith("E"):
        model_series_map["E"].append(m)
    elif s.startswith("L"):
        model_series_map["L"].append(m)
    elif s.startswith("Z"):
        model_series_map["Z"].append(m)

print(f"\n  按宏系列分布:")
all_covered = True
for series, models in model_series_map.items():
    print(f"    {series}: {len(models)} 个型号")
    if len(models) == 0:
        print(f"      ❌ 系列 {series} 完全缺失!")
        all_covered = False

# Check year coverage
years_present = set(r.get("发布年份") for r in rows)
years_expected = set(range(2019, 2026))
print(f"\n  年份覆盖: {sorted(years_present)}")
missing_years = years_expected - years_present
if missing_years:
    print(f"  ❌ 缺失年份: {sorted(missing_years)}")

if missing_in_excel:
    print(f"\n  缺失型号详情:")
    for series, ms in sorted(missing_by_series.items()):
        print(f"    {series} ({len(ms)}): {', '.join(ms[:10])}{'...' if len(ms)>10 else ''}")

if extra_in_excel:
    print(f"\n  额外型号 (不在原始清单中): {len(extra_in_excel)}")
    extras_list = sorted(extra_in_excel)
    for e in extras_list[:10]:
        print(f"    + {e}")
    if len(extras_list) > 10:
        print(f"    ... and {len(extras_list) - 10} more")

# ============ 5. Critical Field Completeness ============
print(f"\n[2/7] 官方硬件参数100%完整")
critical_fields = [
    "CPU型号", "分辨率", "尺寸(英寸)", "刷新率(Hz)", "触摸屏",
    "内存类型", "最大内存", "存储接口", "接口", "指纹识别", "摄像头",
    "集成显卡", "独立显卡", "无线", "官方标称续航(h)", "重量"
]

completeness = {}
total = len(rows)
na_indicators = {"N/A", "N/A", "", None, "无", "None"}

for field in critical_fields:
    missing = sum(1 for r in rows if r.get(field) is None or str(r.get(field)).strip() in na_indicators)
    pct = (total - missing) / total * 100
    completeness[field] = {"present": total - missing, "total": total, "pct": round(pct, 1)}
    status = "❌" if missing > 0 else "✅"
    if missing > 0:
        print(f"  {status} {field}: {total - missing}/{total} ({pct:.1f}%) - {missing} missing")

# ============ 6. Ubuntu / Linux / Price ============
print(f"\n[3/7] Ubuntu Certified认证状态")
ubuntu_status = Counter()
for r in rows:
    v = str(r.get("Ubuntu认证", ""))
    if v in ("是", "yes", "True", "YES"):
        ubuntu_status["是"] += 1
    elif v in ("否", "no", "False", "NO"):
        ubuntu_status["否"] += 1
    else:
        ubuntu_status["未确认/NA"] += 1
print(f"  分布: {dict(ubuntu_status)}")
print(f"  覆盖率: {ubuntu_status['是'] + ubuntu_status['否']}/{total} ({round((ubuntu_status['是']+ubuntu_status['否'])/total*100,1)}%)")

print(f"\n[4/7] Linux兼容性综合评分")
linux_scores = Counter()
for r in rows:
    s = str(r.get("Linux兼容性评分", ""))
    linux_scores[s] += 1
print(f"  分布: {dict(linux_scores)}")

print(f"\n[5/7] 中国二手价格区间")
price_present = sum(1 for r in rows if r.get("二手价格区间(CNY)") and str(r.get("二手价格区间(CNY)")).strip() not in na_indicators)
price_na = total - price_present
print(f"  有价格: {price_present}/{total} | N/A: {price_na}/{total}")
# Sample valid prices
valid_prices = [r.get("二手价格区间(CNY)") for r in rows if r.get("二手价格区间(CNY)") and str(r.get("二手价格区间(CNY)")).strip() not in na_indicators]
print(f"  有效价格样例: {valid_prices[:5]}")

print(f"\n[6/7] Excel条件格式")
cf_rules = []
for cf in ws.conditional_formatting:
    for rule in cf.rules:
        cf_rules.append(f"Range={cf}, Type={rule.type}, Priority={rule.priority}")
print(f"  条件格式规则数: {len(cf_rules)}")
for rl in cf_rules[:6]:
    print(f"    - {rl}")
print(f"  冻结窗格: {ws.freeze_panes}")
print(f"  自动筛选: {ws.auto_filter.ref if ws.auto_filter else 'NONE'}")

print(f"\n[7/7] 数据来源可追溯")
source_present = sum(1 for r in rows if r.get("数据来源URL") and str(r.get("数据来源URL")).strip() not in na_indicators)
source_na = total - source_present
print(f"  有来源URL: {source_present}/{total} | N/A: {source_na}/{total}")
date_present = sum(1 for r in rows if r.get("数据更新日期") and str(r.get("数据更新日期")).strip() not in na_indicators)
print(f"  有数据更新日期: {date_present}/{total}")

# ============ 7. N/A Ratio Analysis ============
print(f"\n{'=' * 70}")
print("N/A RATIO ANALYSIS")
print("=" * 70)

print(f"\n  {'字段':<28} {'N/A':>5} {'%':>6} {'Status'}")
print(f"  {'-'*48}")
high_na_fields = []
for h in headers:
    na_count = sum(1 for r in rows if r.get(h) is None or str(r.get(h)).strip() in na_indicators)
    na_pct = na_count / total * 100
    status = "🔴" if na_pct > 50 else ("🟡" if na_pct > 20 else "🟢")
    if na_pct > 5:  # Only show fields with >5% N/A
        print(f"  {status} {h:<26} {na_count:>5} {na_pct:>5.1f}%")
    if na_pct > 20:
        high_na_fields.append((h, na_count, na_pct))

overall_na = sum(1 for r in rows for h in headers if r.get(h) is None or str(r.get(h)).strip() in na_indicators)
overall_cells = total * len(headers)
overall_na_pct = overall_na / overall_cells * 100
print(f"\n  OVERALL N/A: {overall_na}/{overall_cells} = {overall_na_pct:.1f}%")

# ============ 8. Fabrication Detection ============
print(f"\n{'=' * 70}")
print("FABRICATION DETECTION")
print("=" * 70)

fab_issues = []

# 8a. CPU field: Excel has "N/A" but raw spec has data
cpu_disc = 0
for r in rows:
    model = strip_thinkpad(r.get("型号", ""))
    excel_cpu = str(r.get("CPU型号", ""))
    raw = raw_specs.get(model, {})
    if raw:
        raw_cpu = raw.get("cpu", {})
        if isinstance(raw_cpu, dict):
            raw_has_cpu = bool(raw_cpu.get("family") or raw_cpu.get("models"))
        else:
            raw_has_cpu = bool(raw_cpu)
        if excel_cpu.strip() in na_indicators and raw_has_cpu:
            cpu_disc += 1
            fab_issues.append(f"CPU缺失但Raw有数据: {model}")
print(f"  CPU: Excel=N/A but raw has data: {cpu_disc} models")

# 8b. Resolution field
res_disc = 0
for r in rows:
    model = strip_thinkpad(r.get("型号", ""))
    excel_res = str(r.get("分辨率", ""))
    raw = raw_specs.get(model, {})
    if raw:
        raw_screen = raw.get("screen", {})
        if isinstance(raw_screen, dict):
            raw_has_res = bool(raw_screen.get("resolution_options"))
        else:
            raw_has_res = bool(raw_screen)
        if excel_res.strip() in na_indicators and raw_has_res:
            res_disc += 1
            fab_issues.append(f"分辨率缺失但Raw有数据: {model}")
print(f"  分辨率: Excel=N/A but raw has data: {res_disc} models")

# 8c. Ubuntu certification cross-check
ubuntu_cert_raw = raw_compat.get("ubuntu_cert.json", {})
if ubuntu_cert_raw:
    raw_cert_models = set()
    raw_cert_data = ubuntu_cert_raw.get("models", ubuntu_cert_raw) if isinstance(ubuntu_cert_raw, dict) else ubuntu_cert_raw
    for m in raw_cert_data:
        raw_cert_models.add(strip_thinkpad(m.get("model", "")))

    ubuntu_disc = 0
    for r in rows:
        model = strip_thinkpad(r.get("型号", ""))
        excel_cert = str(r.get("Ubuntu认证", ""))
        if excel_cert in ("是", "yes", "YES"):
            if model not in raw_cert_models:
                ubuntu_disc += 1
                fab_issues.append(f"Ubuntu认证='是'但不在Raw中: {model}")
    print(f"  Ubuntu认证: Excel='是' but not in raw data: {ubuntu_disc} models")

# 8d. Check for impossible/improbable values
improbable = 0
for r in rows:
    size = r.get("尺寸(英寸)")
    if size and str(size).strip() not in na_indicators:
        try:
            sz = float(str(size).replace('"', '').replace('"', '').replace('英寸', '').strip())
            if sz < 10 or sz > 18:
                improbable += 1
                fab_issues.append(f"不合理屏幕尺寸: {r.get('型号')} = {size}")
        except:
            pass

    cpu = str(r.get("CPU型号", ""))
    if cpu.strip() not in na_indicators:
        # Check if CPU looks plausible
        if not re.search(r'(i\d|Ryzen|Core|Intel|AMD|Snapdragon|Apple|Xeon|Qualcomm|MediaTek)', cpu, re.I):
            improbable += 1
            fab_issues.append(f"可疑CPU: {r.get('型号')} = {cpu}")

print(f"  不合理/可疑值: {improbable} occurrences")

# 8e. Check price format
price_format_issues = 0
for r in rows:
    price = str(r.get("二手价格区间(CNY)", ""))
    if price.strip() not in na_indicators:
        if not re.search(r'[¥\d]', price):
            price_format_issues += 1
print(f"  价格格式问题: {price_format_issues}")

# ============ 9. Critical Missing Deep Dive ============
print(f"\n{'=' * 70}")
print("CRITICAL FIELD: 缺失详情")
print("=" * 70)

critical_missing = []
for r in rows:
    model = r.get("型号", "")
    issues = []
    for field in ["CPU型号", "分辨率"]:
        v = str(r.get(field, ""))
        if v.strip() in na_indicators:
            issues.append(field)
    if issues:
        critical_missing.append((model, issues, r.get("系列", ""), r.get("发布年份", "")))

print(f"CPU或分辨率缺失的型号: {len(critical_missing)}")
for model, issues, series, year in critical_missing[:15]:
    print(f"  ❌ {model} | {series} | {year} | {'+'.join(issues)}")
if len(critical_missing) > 15:
    print(f"  ... and {len(critical_missing) - 15} more")

# ============ 10. Raw Data Cross-Reference ============
print(f"\n{'=' * 70}")
print("RAW DATA CROSS-REFERENCE")
print("=" * 70)

# Models that have raw spec data available
excel_with_raw = 0
excel_without_raw = 0
for r in rows:
    model = strip_thinkpad(r.get("型号", ""))
    if model in raw_specs:
        excel_without_raw += 1
        # Check consistency between Excel and raw
        raw = raw_specs[model]
        # Check screen size
        raw_screen = raw.get("screen", {})
        if isinstance(raw_screen, dict) and raw_screen.get("size"):
            excel_size = str(r.get("尺寸(英寸)", ""))
            raw_size = raw_screen["size"]
            if excel_size.strip() not in na_indicators and raw_size.strip() not in ("", "N/A"):
                # Normalize sizes for comparison
                excel_norm = excel_size.replace('"', '').replace('英寸', '').strip()
                raw_norm = raw_size.replace('"', '').replace('英寸', '').strip()
                if excel_norm != raw_norm:
                    fab_issues.append(f"尺寸不一致: {model} Excel={excel_size} Raw={raw_size}")
    else:
        excel_without_raw += 1

print(f"  Excel型号含Raw数据: {excel_with_raw}")
print(f"  Excel型号无Raw数据: {excel_without_raw}")
print(f"  原始规格数据可用率: {excel_with_raw}/{len(rows)} ({excel_with_raw/len(rows)*100:.1f}%)")

# ============ 11. VERDICT ============
print(f"\n{'=' * 70}")
print("FINAL VERDICT")
print("=" * 70)

# Count critical data quality issues
max_na_pct = max(pct for _, _, pct in high_na_fields) if high_na_fields else 0

# Must-have scores
mh_scores = {
    "系列覆盖": 1.0 if len(missing_in_excel) == 0 and all_covered else (0.7 if len(missing_in_excel) <= 5 else 0.3),
    "官方参数完整": 1.0 if len(critical_missing) == 0 else (0.5 if len(critical_missing) <= 10 else 0.0),
    "Ubuntu认证标注": 1.0 if ubuntu_status.get("未确认/NA", 0) < total * 0.3 else 0.7,
    "Linux兼容性评分": 1.0 if len(linux_scores) > 2 else 0.5,
    "二手价格": min(1.0, price_present / total),
    "Excel格式": 1.0 if len(cf_rules) > 0 else 0.0,
    "数据来源可追溯": min(1.0, source_present / total),
}

print(f"\n  Must-Have Scores:")
total_score = 0
for k, v in mh_scores.items():
    total_score += v
    icon = "✅" if v >= 0.8 else ("⚠️" if v >= 0.5 else "❌")
    print(f"  {icon} {k}: {v:.0%}")

avg_score = total_score / len(mh_scores)
print(f"\n  Average Score: {avg_score:.1%}")
print(f"  Critical Missing (CPU/Resolution): {len(critical_missing)} models")
print(f"  Overall N/A: {overall_na_pct:.1f}%")
print(f"  Fabrication Issues: {len(fab_issues)}")

# Verdict logic
if len(fab_issues) > 5:
    verdict = "REJECT"
elif len(critical_missing) > 30:
    verdict = "REJECT"
elif len(critical_missing) > 10:
    verdict = "REJECT - Critical data gaps must be resolved"
elif avg_score < 0.6:
    verdict = "REJECT - Score too low"
elif avg_score < 0.8:
    verdict = "APPROVE WITH CAVEATS"
else:
    verdict = "APPROVE"

# Output summary line
must_have_passed = sum(1 for v in mh_scores.values() if v >= 0.8)
fab_status = f"CLEAN" if len(fab_issues) == 0 else f"{len(fab_issues)} issues"
print(f"\n  >>> OUTPUT: Must Have [{must_have_passed}/{len(mh_scores)}] | N/A Ratio [{overall_na_pct:.1f}%] | Fabrication [{fab_status}] | VERDICT: {verdict} <<<")

# Save report
report_path = ".sisyphus/evidence/task-22-audit.txt"
with open(report_path, "w", encoding="utf-8") as f:
    f.write(f"Task 22: 数据完整性审计报告\n")
    f.write(f"审计时间: {datetime.now().isoformat()}\n")
    f.write(f"{'='*70}\n\n")
    f.write(f"VERDICT: {verdict}\n")
    f.write(f"Must Have Passed: {must_have_passed}/{len(mh_scores)}\n")
    f.write(f"Average Score: {avg_score:.1%}\n")
    f.write(f"Overall N/A Ratio: {overall_na_pct:.1f}%\n")
    f.write(f"Fabrication Issues: {len(fab_issues)}\n")
    f.write(f"Critical Missing (CPU/Resolution): {len(critical_missing)} models\n\n")
    f.write(f"=== Must-Have Scores ===\n")
    for k, v in mh_scores.items():
        f.write(f"  {k}: {v:.0%}\n")
    f.write(f"\n=== Fabrication Issues ===\n")
    for issue in fab_issues:
        f.write(f"  - {issue}\n")
    f.write(f"\n=== Critical Missing Models ===\n")
    for model, issues, series, year in critical_missing:
        f.write(f"  - {model} ({series}, {year}): {'+'.join(issues)}\n")
    f.write(f"\n=== High N/A Fields (>20%) ===\n")
    for h, c, p in high_na_fields:
        f.write(f"  - {h}: {p:.1f}% ({c}/{total})\n")

print(f"\n  报告已保存: {report_path}")
